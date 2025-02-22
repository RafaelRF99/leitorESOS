import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook

EXCEL_PATH = r"\\srv4flabeg\DOCS2\ENGENHARIA\GERENCIAMENTO DE PROGRAMAS\Tabela link.xlsm"
SHEET_NAME = "Rendimento programa"

resultados = []
ultimo_codigo = ""
filtro = ""

def buscar_dados(codigo_barra):
    try:
        wb = load_workbook(EXCEL_PATH, data_only=True)
        aba = wb[SHEET_NAME]
        
        codigo_encontrado = None
        resultados_local = []
        
        for row in aba.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= 6 and str(row[5]).strip() == codigo_barra.strip():
                codigo_encontrado = row[0]
                break
        
        if not codigo_encontrado:
            return None
        
        for row in aba.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= 6 and str(row[0]).strip() == str(codigo_encontrado).strip():
                resultados_local.append(row)
        
        return resultados_local if resultados_local else None
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao ler o arquivo Excel: {e}")
        return None

def filtrar_resultado(resultados, filtro):
    if filtro:
        resultados_filtrados = []
        for row in resultados:
            try:
                # Converte ambos para string e remove espaços extras
                valor_celula = str(row[4]).strip()
                valor_filtro = str(filtro).strip()
                
                # Se os valores forem numéricos, converte para inteiro antes de comparar
                if valor_celula.isdigit() and valor_filtro.isdigit():
                    if int(valor_celula) == int(valor_filtro):
                        resultados_filtrados.append(row)
                else:
                    if valor_filtro.lower() in valor_celula.lower():
                        resultados_filtrados.append(row)
            except Exception as e:
                print(f"Erro ao filtrar: {e}")
        return resultados_filtrados
    return resultados


def exibir_resultado(resultados):
    texto_resultado.config(state="normal")  # Habilita edição para inserir texto
    texto_resultado.delete(1.0, tk.END)  # Limpa o conteúdo

    if resultados:
        col_sizes = [10, 15, 15, 10, 10]
        header = f"{'Linha':^{col_sizes[0]}}{'Código':^{col_sizes[1]}}{'Raio':^{col_sizes[2]}}{'MP':^{col_sizes[3]}}{'REND':^{col_sizes[4]}}\n"
        texto_resultado.insert(tk.END, header, "header")
        texto_resultado.insert(tk.END, "-" * sum(col_sizes) + "\n")

        for i, row in enumerate(resultados):
            linha_formatada = f"{str(row[4]):^{col_sizes[0]}}{str(row[0]):^{col_sizes[1]}}{str(row[1]):^{col_sizes[2]}}{str(row[2]):^{col_sizes[3]}}{str(row[3]):^{col_sizes[4]}}\n"
            tag = "even" if i % 2 == 0 else "odd"
            texto_resultado.insert(tk.END, linha_formatada, tag)
    
    else:
        texto_resultado.insert(tk.END, "Código não encontrado.")

    texto_resultado.config(state="disabled")

def on_enter(event=None):
    global ultimo_codigo, resultados
    codigo = codigo_var.get().strip()
    if len(codigo) != 11:
        return
    
    ultimo_codigo = codigo
    resultados = buscar_dados(codigo)
    if resultados:
        resultados = filtrar_resultado(resultados, filtro)
    exibir_resultado(resultados)
    codigo_var.set("")
    entry_codigo.focus()

def aplicar_filtro():
    global filtro, resultados
    filtro = filtro_var.get().strip()
    if resultados:
        resultados_filtrados = filtrar_resultado(resultados, filtro)
        exibir_resultado(resultados_filtrados)
    elif ultimo_codigo:
        resultados = buscar_dados(ultimo_codigo)
        if resultados:
            resultados = filtrar_resultado(resultados, filtro)
            exibir_resultado(resultados)

def limitar_tamanho(*args):
    texto = codigo_var.get()
    if len(texto) > 11:
        codigo_var.set(texto[:11])
    elif len(texto) == 11:
        root.after(100, on_enter)

root = tk.Tk()
root.title("Busca no Excel")
root.geometry("800x600")

tk.Label(root, text="Digite o código abaixo:", font=("Arial", 12)).pack(pady=10)

codigo_var = tk.StringVar()
codigo_var.trace_add("write", limitar_tamanho)
entry_codigo = tk.Entry(root, font=("Arial", 14), textvariable=codigo_var, width=20, bd=2, relief="solid", justify="center")
entry_codigo.pack(pady=10)
entry_codigo.focus()

frame_filtro = tk.Frame(root)
frame_filtro.pack(pady=10, anchor="w", padx=10)
tk.Label(frame_filtro, text="Linha:", font=("Arial", 12)).pack(side=tk.LEFT)

filtro_var = tk.StringVar()
entry_filtro = tk.Entry(frame_filtro, font=("Arial", 14), textvariable=filtro_var, width=8, bd=2, relief="solid", justify="center")
entry_filtro.pack(side=tk.LEFT)
btn_filtrar = tk.Button(frame_filtro, text="Filtrar", command=aplicar_filtro)
btn_filtrar.pack(side=tk.LEFT, padx=10)

texto_resultado = tk.Text(root, font=("Courier", 12), width=90, height=20, wrap=tk.WORD)
texto_resultado.pack(pady=20)
texto_resultado.tag_configure("even", background="#F0F0F0")
texto_resultado.tag_configure("odd", background="#D0D0D0")
texto_resultado.tag_configure("header", font=("Courier", 12, "bold"))
texto_resultado.config(state="disabled")

tk.Label(root, text="Caso dificuldade, digitar 'help'", font=("Arial", 10)).pack(side=tk.BOTTOM, fill="x", padx=20)

root.mainloop()
